import os
import sys
import re
import numpy as np
import pandas as pd
from dotenv import load_dotenv
from sqlalchemy import create_engine, text
from openpyxl import load_workbook  # New import for formatting

# ------------------------------------------------------------------
# 1) Load DB credentials from env
# ------------------------------------------------------------------
script_dir = os.path.dirname(__file__)
dotenv_path = os.path.join(script_dir, "Credentials.env")
load_dotenv(dotenv_path)

DB_USERNAME = os.getenv("DB_USERNAME")
DB_PASSWORD = os.getenv("DB_PASSWORD")
DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
DB_NAME = os.getenv("DB_NAME")

engine = create_engine(
    f"mysql+pymysql://{DB_USERNAME}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}",
    connect_args={"ssl": {"fake_flag_to_enable": True}}
)

# ------------------------------------------------------------------
# 2) Prepare the user inputs: Ticker & file path
# ------------------------------------------------------------------
ticker = "WPC"  # Update this to the REIT you want to process
print(f"Processing data for Ticker={ticker}...")

root_folder = os.getenv("REIT_RAW_FINANCIALS_PATH")
xlsx_file_name = f"{ticker} Financials.xlsx"
xls_file_name = f"{ticker} Financials.xls"

file_path_xlsx = os.path.join(root_folder, ticker, "Financials", xlsx_file_name)
file_path_xls = os.path.join(root_folder, ticker, "Financials", xls_file_name)

if os.path.exists(file_path_xlsx):
    file_path = file_path_xlsx
elif os.path.exists(file_path_xls):
    file_path = file_path_xls
else:
    print("❌ No .xlsx or .xls file found for this REIT.\n"
          f"Tried:\n  {file_path_xlsx}\n  {file_path_xls}")
    sys.exit(1)

print(f"Using file: {file_path}")

# ------------------------------------------------------------------
# 3) Helper: Create or verify each statement table
#    We'll store data in row-based format with an additional
#    'fiscal_quarter' column. The unique constraint is now
#    (ticker, line_item, fiscal_year, fiscal_quarter).
#    ADDED: excel_row_index INT for preserving row order.
#           is_bold (TINYINT) and display_format (VARCHAR) for cell formatting.
# ------------------------------------------------------------------
def create_reit_table_if_not_exists(table_name):
    create_query = f"""
    CREATE TABLE IF NOT EXISTS {table_name} (
        id INT AUTO_INCREMENT PRIMARY KEY,
        ticker VARCHAR(10) NOT NULL,
        line_item VARCHAR(255) NOT NULL,
        fiscal_year INT NOT NULL,
        fiscal_quarter INT NULL,
        value FLOAT,
        excel_row_index INT,
        is_bold TINYINT DEFAULT 0,
        display_format VARCHAR(50),
        UNIQUE KEY unique_row (ticker, line_item, fiscal_year, fiscal_quarter)
    );
    """
    try:
        with engine.connect() as conn:
            conn.execute(text(create_query))
        print(f"✅ Verified/created table: {table_name}")
    except Exception as e:
        print(f"❌ Error creating table {table_name}: {e}")
        sys.exit(1)

statement_tables = {
    "Income Statement": "reit_income_statement",
    "Balance Sheet": "reit_balance_sheet",
    "Cash Flow": "reit_cash_flow",
    "Industry Specific": "reit_industry_metrics",
}

for sheet_name, table_name in statement_tables.items():
    create_reit_table_if_not_exists(table_name)

# ------------------------------------------------------------------
# 4) Quarter/Year parsing
#    We'll look for a 4-digit year and "Q1","Q2","Q3","Q4" in
#    the column name. If no quarter is found, defaults to None.
# ------------------------------------------------------------------
def parse_year_quarter(col_name: str):
    """
    Example column strings might be:
      "Reclassified\n3 months\nQ1\nMar-31-2015"
      "3 months Q2 Jun-30-2016"
      "12 months Dec-31-2015"
    We'll extract the 4-digit year + quarter (if present).
    Returns (year, quarter) or (None, None) if not found.
    """
    # Find a 4-digit year
    y_match = re.search(r'(19\d{2}|20\d{2})', col_name)
    if not y_match:
        return (None, None)
    year = int(y_match.group(1))

    # Look for "Q1", "Q2", "Q3", or "Q4"
    q_match = re.search(r'(Q[1-4])', col_name, flags=re.IGNORECASE)
    if q_match:
        # e.g. "Q1" -> quarter=1
        quarter_str = q_match.group(1).upper()
        quarter = int(quarter_str[1])
    else:
        quarter = None

    return (year, quarter)

# ------------------------------------------------------------------
# 5) Function to process a single sheet (using openpyxl to preserve formatting)
#    ADDED: preserve original Excel row order via excel_row_index.
#           Captures the bold state and number format.
# ------------------------------------------------------------------
def process_financial_sheet(file_path, sheet_name, ticker):
    table_name = statement_tables[sheet_name]

    # Load workbook with formatting (assumes xlsx file)
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # Get headers from first row
    header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
    headers = [cell.value for cell in header_cells]

    # Identify the line_item column as the first header that does not include a 4-digit year.
    line_item_col = None
    year_cols_indices = []
    for i, header in enumerate(headers):
        if header is None:
            continue
        if re.search(r'(19\d{2}|20\d{2})', str(header)):
            year_cols_indices.append(i)
        else:
            if line_item_col is None:
                line_item_col = i

    if line_item_col is None:
        line_item_col = 0

    data_rows = []
    # Iterate over rows starting from row 2 (the data rows)
    for row in ws.iter_rows(min_row=2, values_only=False):
        # Get the line_item value from the designated column
        line_item = row[line_item_col].value
        if line_item is None:
            continue
        # For each year column, extract the cell value and formatting info
        for i in year_cols_indices:
            cell = row[i]
            cell_value = cell.value
            # Extract formatting: bold and number format
            is_bold = 1 if cell.font and cell.font.bold else 0
            number_format = cell.number_format
            # Determine display_format based on number_format: if it contains "$" or "%" (custom logic)
            if number_format and "$" in number_format:
                display_format = "currency"
            elif number_format and "%" in number_format:
                display_format = "percent"
            else:
                display_format = None

            # Use the header of this column to extract fiscal year and quarter
            raw_column = headers[i]
            year, quarter = parse_year_quarter(str(raw_column))
            if year is None:
                continue

            record = {
                "ticker": ticker,
                "line_item": line_item,
                "fiscal_year": year,
                "fiscal_quarter": quarter,
                "value": cell_value,
                # Use the Excel row number minus 2 (to zero-index data rows; header is row 1)
                "excel_row_index": cell.row - 2,
                "is_bold": is_bold,
                "display_format": display_format
            }
            data_rows.append(record)

    # Convert the list of records into a DataFrame
    df_melted = pd.DataFrame(data_rows)

    # Drop rows missing a fiscal_year (should already be handled by parse_year_quarter)
    df_melted.dropna(subset=["fiscal_year"], inplace=True)
    df_melted["fiscal_year"] = df_melted["fiscal_year"].astype(int)

    # Insert into MySQL
    try:
        df_melted.to_sql(table_name, con=engine, if_exists='append', index=False)
        print(f"✅ Inserted {len(df_melted)} rows into {table_name}")
    except Exception as e:
        print(f"❌ Error inserting into {table_name}: {e}")

# ------------------------------------------------------------------
# 6) Read the Excel’s multiple sheets and process them using openpyxl
# ------------------------------------------------------------------
wanted_sheets = [
    "Income Statement",
    "Balance Sheet",
    "Cash Flow",
    "Industry Specific"
]

print(f"Reading Excel: {file_path} ...")
# For formatting we use openpyxl instead of pd.read_excel.
for sheet in wanted_sheets:
    try:
        wb = load_workbook(file_path, data_only=True)
        if sheet in wb.sheetnames:
            print(f"🔎 Processing sheet '{sheet}'...")
            process_financial_sheet(file_path, sheet, ticker)
        else:
            print(f"⚠️ Sheet '{sheet}' not found; skipping.")
    except Exception as e:
        print(f"❌ Error processing sheet '{sheet}': {e}")

print("✅ Done processing all sheets for ticker:", ticker)
